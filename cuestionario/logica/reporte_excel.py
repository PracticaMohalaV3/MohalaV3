from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from cuestionario.models import Trabajador, Autoevaluacion, EvaluacionJefatura, ResultadoConsolidado, TextosEvaluacion
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO


@login_required
def generar_excel_detalle(request, trabajador_id):
    """Genera el Excel del reporte de evaluación usando openpyxl"""
    if not request.user.is_superuser:
        return redirect('index')
    
    try:
        trabajador = Trabajador.objects.select_related('cargo', 'nivel_jerarquico', 'id_jefe_directo').get(id_trabajador=trabajador_id)
    except Trabajador.DoesNotExist:
        return redirect('seguimiento_admin')

    empresa = trabajador.empresa

    resultados = ResultadoConsolidado.objects.filter(
        trabajador=trabajador
    ).order_by('textos_evaluacion_codigo_excel')

    codigos = list(resultados.values_list('textos_evaluacion_codigo_excel', flat=True).distinct())
    textos_qs = TextosEvaluacion.objects.filter(
        empresa=empresa,
        codigo_excel__in=codigos
    ).select_related('dimension', 'competencia')
    textos_map = {t.codigo_excel: t for t in textos_qs}

    resultados_por_dim = {}
    for r in resultados:
        texto_eval = textos_map.get(r.textos_evaluacion_codigo_excel)
        if not texto_eval:
            continue
        r.texto_eval = texto_eval
        dim_nombre = texto_eval.dimension.nombre_dimension
        if dim_nombre not in resultados_por_dim:
            resultados_por_dim[dim_nombre] = []
        resultados_por_dim[dim_nombre].append(r)

    auto = Autoevaluacion.objects.filter(trabajador=trabajador, estado_finalizacion=True).first()
    jefe = EvaluacionJefatura.objects.filter(trabajador_evaluado=trabajador, estado_finalizacion=True).first()
    
    timestamp_auto = auto.momento_evaluacion.strftime("%d/%m/%Y %H:%M") if auto else "Pendiente"
    timestamp_jefe = jefe.momento_evaluacion.strftime("%d/%m/%Y %H:%M") if jefe else "N/A"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Evaluación"
    
    title_font = Font(name='Arial', size=16, bold=True, color='5e42a6')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='5e42a6', end_color='5e42a6', fill_type='solid')
    info_fill = PatternFill(start_color='f0f0f0', end_color='f0f0f0', fill_type='solid')
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    ws['A1'] = 'Reporte de Evaluación de Desempeño'
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    row = 3
    info_data = [
        ['Colaborador:', f"{trabajador.nombre} {trabajador.apellido_paterno} {trabajador.apellido_materno}"],
        ['Empresa:', trabajador.empresa.nombre_empresa],
        ['Cargo:', trabajador.cargo.nombre_cargo],
        ['Nivel:', trabajador.nivel_jerarquico.nombre_nivel_jerarquico],
        ['Jefatura Directa:', f"{trabajador.id_jefe_directo.nombre} {trabajador.id_jefe_directo.apellido_paterno} {trabajador.id_jefe_directo.apellido_materno}" if trabajador.id_jefe_directo else "N/A"],
        ['Autoevaluación finalizada:', timestamp_auto],
        ['Evaluación Jefatura finalizada:', timestamp_jefe]
    ]
    
    for label, value in info_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = info_fill
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1

    dim_colors = ['51ff85', '2196F3']
    headers = ['Código', 'Competencia / Indicador', 'AutoEv', 'Ev. Jefe', 'Diferencia']

    for dim_idx, (dim_nombre, lista) in enumerate(resultados_por_dim.items()):
        dim_color = dim_colors[dim_idx] if dim_idx < len(dim_colors) else '5e42a6'

        row += 2
        ws[f'A{row}'] = f'Dimensión: {dim_nombre}'
        ws[f'A{row}'].font = Font(size=12, bold=True, color=dim_color)
        ws.merge_cells(f'A{row}:E{row}')

        row += 1
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        row += 1
        for r in lista:
            ws[f'A{row}'] = r.texto_eval.codigo_excel
            ws[f'B{row}'] = r.texto_eval.competencia.nombre_competencia
            ws[f'C{row}'] = r.puntaje_autoev
            ws[f'D{row}'] = r.puntaje_jefe if r.puntaje_jefe > 0 else 'N/A'
            ws[f'E{row}'] = int(r.diferencia)

            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].border = border
                ws[f'{col}{row}'].alignment = Alignment(horizontal='center' if col != 'B' else 'left', vertical='center')

            if r.diferencia > 0:
                ws[f'E{row}'].fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
                ws[f'E{row}'].font = Font(color='155724', bold=True)
            elif r.diferencia < 0:
                ws[f'E{row}'].fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
                ws[f'E{row}'].font = Font(color='721c24', bold=True)

            row += 1
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_{trabajador.nombre}_{trabajador.apellido_paterno}.xlsx"'
    
    return response